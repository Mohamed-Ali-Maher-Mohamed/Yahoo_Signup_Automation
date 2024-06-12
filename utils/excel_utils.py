# utils/excel_utils.py
import openpyxl
import os
from openpyxl.styles import PatternFill

def read_excel(sheet_name):
    # Construct the absolute path to the Excel file

    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, 'Mohamed_data.xlsx')

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Assuming the first row is the header
    headers = data[0]
    rows = data[1:]

    return [dict(zip(headers, row)) for row in rows]


def write_excel(sheet_name, data):
    # Get the directory of the current script
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    excel_path = os.path.join(base_dir, 'Mohamed_data.xlsx')

    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Find the status and reason columns
    status_col = reason_col = created_email_col = None
    for idx, cell in enumerate(sheet[1]):
        if cell.value == 'Status':
            status_col = idx + 1
        elif cell.value == 'Reason':
            reason_col = idx + 1
        elif cell.value == 'created_email':
            created_email_col = idx + 1

    # Clear the status and reason columns
    if status_col:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=status_col, max_col=status_col):
            for cell in row:
                cell.value = None
                cell.fill = PatternFill(fill_type=None)  # Clear the highlight

    if reason_col:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=reason_col, max_col=reason_col):
            for cell in row:
                cell.value = None

    if created_email_col:
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=created_email_col, max_col=created_email_col):
            for cell in row:
                cell.value = None

    # Write the new data
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row_idx, record in enumerate(data, start=2):
        for col_idx, (header, value) in enumerate(record.items(), start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            if header == 'Status':
                if value == 'Passed':
                    cell.fill = green_fill
                elif value == 'Failed':
                    cell.fill = red_fill

    workbook.save(excel_path)